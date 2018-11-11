import sys
import io
from bs4 import BeautifulSoup
import os
import openpyxl
import re

sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding = 'utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding = 'utf-8')

# 엑셀파일 열기
wb = openpyxl.load_workbook('C:/open.xlsx')
ws = wb.active
endpoint1 = 0
endpoint2 = 0
mapdata = []
namelist = []
contactlist = []
for number in range(1, 17):



    fulfilename = os.path.join("C:/Users/gjdigj145/OneDrive/일/스타벅스/HTML/", "C:/Users/gjdigj145/OneDrive/일/스타벅스/HTML/" + str(number) + ".html")

    html = open(fulfilename, 'r', encoding='UTF8')
    soup = BeautifulSoup(html, 'html.parser')
    #print(soup.find_all(class_='quickResultLstCon'))

    Big_target = soup.select_one("#mCSB_3_container > ul")

    for j in Big_target.find_all("strong"):
        namelist.append(j.get_text())

    for j in Big_target.select("p > a"):
        contactlist.append(j.get_text())

    for a in soup("a"):
        soup.a.decompose()
    Big_target = soup.select_one("#mCSB_3_container > ul")

    for j in Big_target.select("p"):
        #mapdata.append(j.get_text().split())
        mapdata.append(j.get_text())
    print(mapdata[1])

    print("매장이름 개수 >> ", len(namelist))
    print("연락처 개수 >> ", len(contactlist))
    print("주소 개수 >> ", len(mapdata))


for index in range(len(mapdata)):
    ws.cell(row=index+1, column=1).value = mapdata[index]

    pattern0 = re.compile(r"도$")
    pattern1 = re.compile(r"시$")
    pattern2 = re.compile(r"구$")
    pattern3 = re.compile(r"군$")
    pointer = 0
    last_string = []

    for index in range(len(namelist)):
        # 이름쓰기
        ws.cell(row=index+1+endpoint1, column=1).value = namelist[index]

        #연락처 쓰기
        ws.cell(row=index+1+endpoint1, column=9).value = contactlist[index]

    endpoint1 += len(namelist)

    for i in range(len(mapdata)):
        last_string = []
        pointer = 0
        for j in range(len(mapdata[i])):

            condition0 = re.search(pattern0, mapdata[i][j])
            condition1 = re.search(pattern1, mapdata[i][j])
            condition2 = re.search(pattern2, mapdata[i][j])
            condition3 = re.search(pattern3, mapdata[i][j])

            if condition0:
                pointer += 1#포인터의 수 만큼 앞의거를 지우는거임
                ws.cell(row=i+1+endpoint2, column=2).value = mapdata[i][j]

            if condition1:
                pointer += 1#특별시 같은 것은 도가 없기 때문에 도를 안지워도 됨
                ws.cell(row=i+1+endpoint2, column=3).value = mapdata[i][j]

            if condition2 or condition3:
                pointer += 1
                ws.cell(row=i+1+endpoint2, column=4).value = mapdata[i][j]
        ws.cell(row=i+1+endpoint2, column=4).value = ' '.join(mapdata[i][pointer:])
    endpoint2 += len(mapdata)




    print(number,"번 작업끝!")
    #엑셀 파일 저장
wb.save("C:/starbucks_mapdata.xlsx")
wb.close()
