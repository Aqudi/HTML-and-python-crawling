# -*- coding: utf-8 -*-

from bs4 import BeautifulSoup
import os
import sys
import io
import openpyxl
import re

sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding = 'utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding = 'utf-8')

with open(os.getcwd() + "/working_htmls/Macdonald.html", mode='r', encoding='utf8') as f:
    html = f.read()
nameList = []
phoneList = []
addressList = []

soup = BeautifulSoup(html, "html.parser")
name = soup.find_all("a", title=True, attrs={'data-id':"name"})
phone = soup.find_all("span", class_='phone', attrs={'data-id':'phone'})
address = soup.find_all("p", class_="newAddress", attrs={'data-id':'newaddr'})
matching = {}
matching2 = {}

for i, j in enumerate(name):
    nameList.append(j['title'])
for i in phone:
    phoneList.append(i.string)
for i in address:
    if i.string is not None:
        addressList.append(i.string)

    else:
        addressList.append("No content")

#test code
for i in range(len(nameList)):
    print("매장명 :{}, 연락처 :{}, 주소 :{}".format(nameList[i], phoneList[i], addressList[i]))

for i in range(len(phoneList)):
    matching.update({nameList[i]:phoneList[i]})
for i in range(len(addressList)):
    matching2.update({phoneList[i]:addressList[i]})
#print(matching)

#엑셀데이터쓰기
wb = openpyxl.load_workbook('C:/open.xlsx')
ws = wb.active

#이름, 연락처, 원주소 쓰기
for line, name in enumerate(nameList):
    ws.cell(row=line + 1, column=1).value = name
for line, phone in enumerate(phoneList):
    ws.cell(row=line + 1, column=2).value = phone
for line, address in enumerate(addressList):
    ws.cell(row=line + 1, column=8).value = address

    #addressList.append(ws.cell(row=line + 1, column=8).value)

pattern1 = re.compile(r"시$")
pattern2 = re.compile(r"구$")
pattern3 = re.compile(r"군$")


for line, comp1 in enumerate(addressList):
    pointer = 0
    last_string = []
    comp1 = comp1.split(" ")

    # 도(시)통째로 분류하기 *수작업 필요(엑셀필터)
    ws.cell(row=line + 1, column=3).value = comp1[0]
    comp1 = comp1[1:]
    #print(comp1)
    # 시, 구, 군 분류하기
    for pointer, comp2 in enumerate(comp1):
        condition1 = re.search(pattern1, comp2)
        condition2 = re.search(pattern2, comp2)
        condition3 = re.search(pattern3, comp2)

        if condition1:
            ws.cell(row=line + 1, column=4).value = comp2
            comp1 = comp1[1:]
        elif condition2 or condition3:
            ws.cell(row=line + 1, column=5).value = comp2
            comp1 = comp1[1:]
    ws.cell(row=line + 1, column=6).value = ' '.join(comp1[:])

wb.save("C:/Users/gjdigj145/Macdonald.xlsx")
wb.close()
