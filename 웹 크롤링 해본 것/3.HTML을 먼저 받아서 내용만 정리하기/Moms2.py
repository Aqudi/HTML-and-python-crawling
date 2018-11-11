# -*- coding: utf-8 -*-
import sys
import io
from bs4 import BeautifulSoup
import os
import re
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding = 'utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding = 'utf-8')

nameList = []
phoneList = []
addressList = []
num = 0
with open(os.getcwd()+"/Moms_htmls/html.html", mode='r', encoding='utf8') as f:
    html = f.read()
    soup = BeautifulSoup(html, 'html.parser')

    phoneSource = soup.find_all("td", class_="td_Left")
    for i in phoneSource:
        phone = i.find_next_sibling().string
        phoneList.append(phone)

    nameAddressSourse = soup.find_all("a")
    for i in nameAddressSourse:
        if nameAddressSourse.index(i)%3 == 1:
            addressList.append(i.string)
            #print(i)
        elif nameAddressSourse.index(i)%3 == 0:
            nameList.append(i.string)
            #print(i)
        else:
            num+=1
print(nameList)
print(addressList)
print(phoneList)


#엑셀데이터쓰기
wb = openpyxl.load_workbook('C:/open.xlsx')
ws = wb.active

for line in range(len(nameList)):
    #이름, 연락처 쓰기
    ws.cell(row=line + 1, column=1).value = nameList[line]
    ws.cell(row=line + 1, column=2).value = phoneList[line]
    ws.cell(row=line + 1, column=8).value = addressList[line]

pattern1 = re.compile(r"시$")
pattern2 = re.compile(r"구$")
pattern3 = re.compile(r"군$")


for line, comp1 in enumerate(addressList):
    pointer = 0
    last_string = []
    print(comp1)
    comp1 = comp1.split(" ")
    print(comp1)

    # 도, 광역시 쓰기
    if re.search(pattern1, comp1[1]):
        #1번항목이 ~도에 속한다.
        ws.cell(row=line + 1, column=3).value = comp1[0] + "도"
        pointer += 1
    if re.search(pattern2, comp1[1]) or re.search(pattern3, comp1[1]):
        #1번 항목이 ~시에 속한다.
        ws.cell(row=line + 1, column=4).value = comp1[0] + "시"
        pointer += 1

    for comp2 in comp1[1:]:
        condition1 = re.search(pattern1, comp2)
        condition2 = re.search(pattern2, comp2)
        condition3 = re.search(pattern3, comp2)

        if condition1:
            ws.cell(row=line + 1, column=4).value = comp2
            pointer += 1
        elif condition2 or condition3:
            ws.cell(row=line + 1, column=5).value = comp2
            pointer += 1
    ws.cell(row=line + 1, column=6).value = ' '.join(comp1[pointer:])

wb.save("C:/Users/gjdigj145/Mom's_touch.xlsx")
wb.close()

