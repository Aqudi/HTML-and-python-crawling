# -*- coding: utf-8 -*-
import sys
import io
from bs4 import BeautifulSoup
import os
import re
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding = 'utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding = 'utf-8')

nameList = []
phoneList = []
addressList = []
num = 0
with open(os.getcwd()+"/Moms_htmls/html.html", mode='r', encoding='utf8') as f:
    html = f.read()
    soup = BeautifulSoup(html, 'html.parser')

    phoneSource = soup.find_all("td", class_="td_Left")
    for i in phoneSource:
        phone = i.find_next_sibling().string
        phoneList.append(phone)

    nameAddressSourse = soup.find_all("a")
    for i in nameAddressSourse:
        if nameAddressSourse.index(i)%3 == 1:
            addressList.append(i.string)
            #print(i)
        elif nameAddressSourse.index(i)%3 == 0:
            nameList.append(i.string)
            #print(i)
        else:
            num+=1
print(nameList)
print(addressList)
print(phoneList)


#엑셀데이터쓰기
wb = openpyxl.load_workbook('C:/open.xlsx')
ws = wb.active

for line in range(len(nameList)):
    #이름, 연락처 쓰기
    ws.cell(row=line + 1, column=1).value = nameList[line]
    ws.cell(row=line + 1, column=2).value = phoneList[line]
    ws.cell(row=line + 1, column=8).value = addressList[line]

wb.save("C:/Users/gjdigj145/Mom's_touch_name.xlsx")
wb.close()

