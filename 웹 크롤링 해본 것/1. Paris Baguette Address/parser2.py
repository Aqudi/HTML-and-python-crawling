from bs4 import BeautifulSoup
import urllib.request as requester
import openpyxl
import string
import re
import sys
import io


sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding = 'utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding = 'utf-8')

base_url= 'https://www.paris.co.kr/store/store_list.jsp?page='
contact = []
name = []
mapdata = []
flag = 1

for i in range(1,448):
    #기본세팅
    url = base_url + str(i)
    res =requester.urlopen(url).read()
    soup = BeautifulSoup(res, "html.parser")

#매장 이름과 연락처를 파싱
    greatin = soup.find_all("td")

    print("\n", i, " page 파싱완료")
    for how_many, contents in enumerate(greatin):
        if contents is not None and contents.string is not None:
            if flag == 1:
                name.append(contents.string)
                flag = 2
            elif flag == 2:
                contact.append(contents.string)
                flag = 1
    print("\n", i, " page 끝냄")

#매장의 주소만 파싱
    greatin2 = soup.find_all("a", class_="_mapOpen")
    print("\n", i, " page 파싱완료")

    for how_many, contents in enumerate(greatin2):
        if contents is not None and contents.string is not None:
            mapdata.append(contents.string.split()[1:])
    print("\n", i, " page 끝냄\n")



# 엑셀파일 열기
wb = openpyxl.load_workbook('C:/score.xlsx')

ws = wb.active

pattern1 = re.compile(r"시$")
pattern2 = re.compile(r"구$")
pattern3 = re.compile(r"군$")
pattern4 = re.compile(r"\)$")
pointer = 0
last_string = []


for index in range(len(contact)):
    # 이름쓰기
    ws.cell(row=index+1, column=1).value = name[index]
    #연락처쓰기
    ws.cell(row=index+1, column=9).value = contact[index]

for i in range(len(mapdata)):
    last_string = []
    pointer = 0
    for j in range(len(mapdata[i])):

        condition1 = re.search(pattern1, mapdata[i][j])
        condition2 = re.search(pattern2, mapdata[i][j])
        condition3 = re.search(pattern3, mapdata[i][j])
        condition4 = re.search(pattern4, mapdata[i][j])

        if condition1:
            pointer = 1
            ws.cell(row=i+1, column=3).value = mapdata[i][j]
        if condition2 or condition3:
            if pointer == 0:
                pointer = 1
            else:
                pointer = 2
            ws.cell(row=i+1, column=4).value = mapdata[i][j]
            #print("구, 군 >> ", mapdata[i][j])
        if condition4:
            last_string = []
            for k in range(pointer ,mapdata[i].index(mapdata[i][j])+1):
                last_string.append(mapdata[i][k])
            last_string = ' '.join(last_string)
            ws.cell(row=i+1, column=5).value = last_string
            #print(last_string)
            last_string = []
            for k in range(mapdata[i].index(mapdata[i][j])+2, len(mapdata[i])):
                last_string.append(mapdata[i][k])
            last_string = ' '.join(last_string)
            ws.cell(row=i+1, column=6).value = last_string
            #print(last_string)

print("작업끝!")
#엑셀 파일 저장
wb.save("C:/totalresult.xlsx")
wb.close()
