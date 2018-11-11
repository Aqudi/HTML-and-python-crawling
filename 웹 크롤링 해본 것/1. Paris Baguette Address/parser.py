from bs4 import BeautifulSoup
import urllib.request as requester
import openpyxl
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding = 'utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding = 'utf-8')

base_url= 'https://www.paris.co.kr/store/store_list.jsp?page='
contact = []
name = []
real_information_map = []
flag = 1

for i in range(1,448):
    #기본세팅
    url = base_url + str(i)
    res =requester.urlopen(url).read()
    soup = BeautifulSoup(res, "html.parser")

    #연락처와 상가명을 함께하는 리스트를 만들었다
    greatin = soup.find_all("td")

    print("\n", i, " page 파싱완료")
    for how_many, contents in enumerate(greatin):
        if contents is not None and contents.string is not None:
            if flag == 1:
                name.append(contents.string)
                flag = 2
            elif flag == 2:
                contact.append(contents.string)
                flag = 1
    print("\n", i, " page 끝냄")

# 엑셀파일 열기
wb = openpyxl.load_workbook('C:/score.xlsx')

ws = wb.active

for index in range(1, len(contact)):
    # 이름쓰기
    ws.cell(row=index, column=1).value = name[index]
    #연락처쓰기
    ws.cell(row=index, column=9).value = contact[index]


    print(index / len(contact) * 100, "%")

#엑셀 파일 저장
wb.save("C:/contact_and_name.xlsx")
wb.close()
